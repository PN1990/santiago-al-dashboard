#!/usr/bin/env python3
"""
Bot para descarregar reservas da Talkguest e importar para o dashboard
(Cloudflare Worker + D1). Corre automaticamente via GitHub Actions.

Fluxo:
  1. Login em app.talkguest.com (email + password)
  2. Se aparecer 2FA, lê o código do Gmail via IMAP e submete
  3. Navega para Reservas -> Lista de Reservas
  4. Ações -> Exportar -> descarrega o XLSX
  5. Lê o XLSX (openpyxl), mapeia/normaliza para o formato da app
  6. Envia para o Worker (/api/bot/import) preservando os campos manuais

Secrets necessários (GitHub Actions):
  TALKGUEST_EMAIL      email de login na Talkguest
  TALKGUEST_PASSWORD   password da Talkguest
  API_URL              URL do Worker (ex: https://santiago-al-dashboard-api.pnstays.workers.dev)
  BOT_SECRET           token partilhado com o Worker
  MAIL_2FA_ADDRESS     caixa de correio que RECEBE o código 2FA da Talkguest
  MAIL_2FA_PASSWORD    App Password / password IMAP dessa caixa (só se houver 2FA)
  MAIL_2FA_IMAP_HOST   (opcional) servidor IMAP; default imap.gmail.com
                       (Outlook: outlook.office365.com · iCloud: imap.mail.me.com)
  MAIL_2FA_FOLDER      (opcional) pasta/label IMAP a ler; default INBOX
                       (define para uma label dedicada se criares um filtro)

NOTA: os seletores de login/2FA/exportação são "best-effort" porque foram
escritos sem acesso ao DOM real da Talkguest. A primeira execução gera
screenshots de debug (/tmp/tg_*.png, publicados como artifact) para afinar.
"""

import os
import re
import time
import random
import glob
import tempfile
import imaplib
import email
from email.utils import parsedate_to_datetime
from datetime import datetime, timezone, date

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException
import openpyxl
import requests

# ── Configuração (via GitHub Secrets) ─────────────────────────────────────────
TALKGUEST_EMAIL    = os.environ["TALKGUEST_EMAIL"]
TALKGUEST_PASSWORD = os.environ["TALKGUEST_PASSWORD"]
API_URL            = os.environ["API_URL"]
BOT_SECRET         = os.environ["BOT_SECRET"]
MAIL_2FA_ADDRESS   = os.environ.get("MAIL_2FA_ADDRESS", "")
MAIL_2FA_PASSWORD  = os.environ.get("MAIL_2FA_PASSWORD", "")
MAIL_2FA_IMAP_HOST = os.environ.get("MAIL_2FA_IMAP_HOST", "").strip() or "imap.gmail.com"
# Pasta/label IMAP a ler. Default INBOX; se usares um filtro com label no Gmail
# (ex: "talkguest-2fa"), define MAIL_2FA_FOLDER com esse nome para leitura robusta.
MAIL_2FA_FOLDER    = os.environ.get("MAIL_2FA_FOLDER", "").strip() or "INBOX"

# ── Parâmetros afináveis ──────────────────────────────────────────────────────
LOGIN_URL          = os.environ.get("TALKGUEST_LOGIN_URL", "https://app.talkguest.com/Theme_UI/Login.aspx")
# Remetente(s) esperado(s) do email de 2FA (match parcial, minúsculas)
TWOFA_FROM_HINTS   = ["talkguest"]
# Regex do código 2FA no corpo do email (default: 6 dígitos)
TWOFA_CODE_REGEX   = re.compile(r"\b(\d{6})\b")
# Palavras que indicam que o email é mesmo de verificação (preferência no match)
TWOFA_KEYWORDS     = ["código", "codigo", "code", "verificação", "verificacao",
                      "verification", "autenticação", "autenticacao", "one-time",
                      "one time", "otp", "2fa"]

# ── Helpers ───────────────────────────────────────────────────────────────────
def esperar(min_s=1.5, max_s=3.5):
    time.sleep(random.uniform(min_s, max_s))

def log(msg):
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}", flush=True)

# ── Selenium ──────────────────────────────────────────────────────────────────
def criar_driver():
    opts = Options()
    opts.add_argument("--headless=new")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--window-size=1400,1000")
    opts.add_argument("--lang=pt-PT")  # forçar a Talkguest a mostrar-se em português
    opts.add_argument("--user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36")
    download_dir = tempfile.mkdtemp()
    prefs = {
        "download.default_directory": download_dir,
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing.enabled": True,
        "intl.accept_languages": "pt-PT,pt",
    }
    opts.add_experimental_option("prefs", prefs)
    driver = webdriver.Chrome(options=opts)
    return driver, download_dir

def _digitar(campo, texto):
    campo.clear()
    for ch in texto:
        campo.send_keys(ch)
        time.sleep(random.uniform(0.04, 0.13))

def clicar_por_texto(driver, texto, timeout=12):
    """Clica no primeiro elemento clicável que contém `texto`."""
    xpaths = [
        f"//*[normalize-space(text())='{texto}']",
        f"//button[contains(normalize-space(.),'{texto}')]",
        f"//a[contains(normalize-space(.),'{texto}')]",
        f"//*[@title='{texto}']",
        f"//*[contains(normalize-space(text()),'{texto}')]",
    ]
    for xpath in xpaths:
        try:
            el = WebDriverWait(driver, max(2, timeout // len(xpaths))).until(
                EC.element_to_be_clickable((By.XPATH, xpath)))
            esperar(0.3, 0.8)
            try:
                el.click()
            except Exception:
                driver.execute_script("arguments[0].click();", el)
            log(f"Clicado '{texto}' via {xpath}")
            return True
        except Exception:
            continue
    # último recurso: JS por textContent exato
    ok = driver.execute_script(
        """
        const alvo = arguments[0];
        const els = document.querySelectorAll('button, a, div, span, li');
        for (const el of els) {
            if (el.textContent.trim() === alvo) { el.click(); return true; }
        }
        return false;
        """, texto)
    if ok:
        log(f"Clicado '{texto}' via JS")
        return True
    return False

# ── Login + 2FA ───────────────────────────────────────────────────────────────
# Seletor abrangente para o campo de email/utilizador (type=text como último recurso)
EMAIL_SELECTOR = ("input[type='email'], input[name='email'], input[name='username'], "
                  "input[id*='email' i], input[name*='email' i], input[placeholder*='email' i], "
                  "input[placeholder*='utilizador' i], input[type='text']")

def _log_inputs(driver, contexto):
    """Lista os campos <input> da página — essencial para afinar seletores."""
    try:
        infos = driver.execute_script(
            "return Array.prototype.slice.call(document.querySelectorAll('input'))"
            ".map(function(i){return {type:i.type,name:i.name,id:i.id,ph:i.placeholder,ac:i.autocomplete};});")
        log(f"Inputs em [{contexto}] ({len(infos)}): {infos}")
        frames = driver.find_elements(By.TAG_NAME, "iframe")
        if frames:
            log(f"⚠️ A página tem {len(frames)} iframe(s) — o formulário pode estar lá dentro.")
    except Exception as e:
        log(f"Não consegui listar inputs: {e}")

def _tem_texto(driver, texto):
    try:
        return texto.lower() in driver.page_source.lower()
    except Exception:
        return False

def fazer_login(driver):
    log(f"A abrir login da Talkguest: {LOGIN_URL}")
    driver.get(LOGIN_URL)
    esperar(3, 5)
    driver.save_screenshot("/tmp/tg_0_login.png")
    log(f"URL: {driver.current_url} | título: {driver.title}")
    _log_inputs(driver, "página de login")
    wait = WebDriverWait(driver, 25)

    log("A localizar campo de email...")
    try:
        campo_email = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, EMAIL_SELECTOR)))
    except TimeoutException:
        driver.save_screenshot("/tmp/tg_0b_sem_campo_email.png")
        log("HTML (excerto): " + driver.page_source[:1500])
        raise Exception("Campo de email não encontrado na página de login. Ver tg_0_login.png e o log de inputs.")
    _digitar(campo_email, TALKGUEST_EMAIL)
    esperar(0.4, 1)

    log("A preencher password...")
    campos_pass = driver.find_elements(By.CSS_SELECTOR, "input[type='password']")
    if not campos_pass:
        # login em 2 passos: submeter email primeiro, password aparece a seguir
        driver.save_screenshot("/tmp/tg_0c_sem_password.png")
        if not clicar_por_texto(driver, "Continuar") and not clicar_por_texto(driver, "Seguinte"):
            campo_email.send_keys(Keys.RETURN)
        esperar(2, 3)
        campos_pass = [WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "input[type='password']")))]
    _digitar(campos_pass[0], TALKGUEST_PASSWORD)
    esperar(0.5, 1.2)
    driver.save_screenshot("/tmp/tg_1_login.png")

    log("A submeter login...")
    # Clicar no BOTÃO de submit real (não no texto "Login", que também é o título).
    submetido = False
    for sel in ["input[type='submit']", "button[type='submit']",
                "input[type='button'][value*='ogin' i]", "form button"]:
        els = driver.find_elements(By.CSS_SELECTOR, sel)
        if els:
            try:
                driver.execute_script("arguments[0].click();", els[-1])
                log(f"Login submetido via {sel}")
                submetido = True
                break
            except Exception:
                continue
    if not submetido:
        campos_pass[0].send_keys(Keys.RETURN)
    esperar(5, 7)
    driver.save_screenshot("/tmp/tg_2_apos_login.png")
    log(f"URL após login: {driver.current_url}")

    # ── 2FA ──
    if precisa_2fa(driver):
        tratar_2fa(driver)

    if "login" in driver.current_url.lower() or precisa_2fa(driver):
        driver.save_screenshot("/tmp/tg_erro_final.png")
        log("HTML (excerto): " + driver.page_source[:1000])
        raise Exception("Login/2FA falhou. Ver screenshots de debug.")
    log(f"Login OK -- URL: {driver.current_url}")

def precisa_2fa(driver):
    """Deteta o ecrã de autenticação multi-fator (escolha de método ou código)."""
    try:
        txt = driver.page_source.lower()
        marcadores = ["multi-fator", "multifator", "multi-factor", "enviar código", "enviar codigo",
                      "send code", "código de verificação", "codigo de verificacao",
                      "verification code", "confirmar a sua identidade", "confirm your identity",
                      "autenticação", "authentication", "two-factor", "two factor", "2fa"]
        if any(m in txt for m in marcadores):
            return True
        campos = driver.find_elements(
            By.CSS_SELECTOR,
            "input[name*='code' i], input[name*='otp' i], "
            "input[autocomplete='one-time-code'], input[inputmode='numeric']")
        return bool(campos)
    except Exception:
        return False

def tratar_2fa(driver):
    """Fluxo 2FA da Talkguest: escolher 'Email' -> 'Enviar código' -> inserir código."""
    log("Ecrã de 2FA detetado.")
    driver.save_screenshot("/tmp/tg_3_2fa_metodo.png")
    inicio = datetime.now(timezone.utc)

    # Passo A: ecrã de escolha de método (SMS / Email / Auth App) + botão "Enviar código"
    if _tem_texto(driver, "Enviar código") or _tem_texto(driver, "Enviar codigo") or _tem_texto(driver, "Send code"):
        log("A escolher método 'Email'...")
        _escolher_email_2fa(driver)
        esperar(0.5, 1.2)
        driver.save_screenshot("/tmp/tg_3b_email_escolhido.png")
        inicio = datetime.now(timezone.utc)
        log("A clicar 'Enviar código'...")
        if not clicar_por_texto(driver, "Enviar código") and not clicar_por_texto(driver, "Enviar codigo"):
            clicar_por_texto(driver, "Send code")
        esperar(5, 7)  # dar tempo ao email de chegar
        driver.save_screenshot("/tmp/tg_3c_codigo_enviado.png")

    # Passo B: ler o código do email e inseri-lo
    log("A ler código do email...")
    codigo = ler_codigo_2fa(inicio)
    log(f"Código 2FA obtido: {codigo}")
    esperar(0.5, 1)
    inserir_codigo_2fa(driver, codigo)
    esperar(3, 5)
    driver.save_screenshot("/tmp/tg_4_apos_2fa.png")

def _escolher_email_2fa(driver):
    """Seleciona a opção 'Email' no ecrã de escolha de método 2FA."""
    ok = driver.execute_script("""
        function clickRadioFor(el){
            var r = el.querySelector && el.querySelector('input[type=radio]');
            if(!r){ var p = el.closest('label,li,div'); if(p){ r = p.querySelector('input[type=radio]'); } }
            (r || el).click();
        }
        var els = Array.prototype.slice.call(document.querySelectorAll('label,div,span,li,button'));
        for (var i=0;i<els.length;i++){
            if (els[i].textContent.trim().toLowerCase() === 'email'){ clickRadioFor(els[i]); return true; }
        }
        var radios = document.querySelectorAll('input[type=radio]');
        if (radios.length >= 2){ radios[1].click(); return true; }  // ordem esperada: SMS, Email, Auth App
        if (radios.length === 1){ radios[0].click(); return true; }
        return false;
    """)
    log(f"Escolha de 'Email' no 2FA: {'ok' if ok else 'rádio não encontrado'}")
    return ok

def inserir_codigo_2fa(driver, codigo):
    campos = driver.find_elements(
        By.CSS_SELECTOR,
        "input[name*='code' i], input[name*='otp' i], input[name*='token' i], "
        "input[autocomplete='one-time-code'], input[inputmode='numeric']")
    if not campos:
        campos = driver.find_elements(By.CSS_SELECTOR, "input[type='text'], input:not([type])")
    if not campos:
        driver.save_screenshot("/tmp/tg_erro_codigo.png")
        raise Exception("Campo para inserir o código 2FA não encontrado.")
    if len(campos) >= len(codigo) and len(campos) >= 4:
        # um input por dígito
        for c, ch in zip(campos, codigo):
            c.send_keys(ch)
            time.sleep(random.uniform(0.05, 0.15))
    else:
        _digitar(campos[0], codigo)
    esperar(0.5, 1.2)
    for txt in ["Confirmar", "Verificar", "Validar", "Entrar", "Confirm", "Verify", "Submit", "Continue"]:
        if clicar_por_texto(driver, txt, timeout=4):
            return
    campos[0].send_keys(Keys.RETURN)

def _texto_email(msg):
    partes = []
    if msg.is_multipart():
        for part in msg.walk():
            ct = part.get_content_type()
            if ct in ("text/plain", "text/html"):
                try:
                    partes.append(part.get_payload(decode=True).decode(
                        part.get_content_charset() or "utf-8", errors="ignore"))
                except Exception:
                    pass
    else:
        try:
            partes.append(msg.get_payload(decode=True).decode(
                msg.get_content_charset() or "utf-8", errors="ignore"))
        except Exception:
            pass
    return "\n".join(partes)

def ler_codigo_2fa(desde_utc, timeout=150):
    """Procura no Gmail (IMAP) o código de 2FA mais recente da Talkguest."""
    if not MAIL_2FA_ADDRESS or not MAIL_2FA_PASSWORD:
        raise Exception("MAIL_2FA_ADDRESS/MAIL_2FA_PASSWORD não definidos, mas a Talkguest pediu 2FA.")
    deadline = time.time() + timeout
    margem = 60  # segundos de tolerância antes do início do login
    fallback = None  # código de um email da Talkguest sem palavras de verificação
    while time.time() < deadline:
        try:
            M = imaplib.IMAP4_SSL(MAIL_2FA_IMAP_HOST)
            M.login(MAIL_2FA_ADDRESS, MAIL_2FA_PASSWORD)
            M.select(f'"{MAIL_2FA_FOLDER}"')
            typ, data = M.search(None, "ALL")
            ids = data[0].split()
            for msg_id in reversed(ids[-25:]):
                typ, msg_data = M.fetch(msg_id, "(RFC822)")
                msg = email.message_from_bytes(msg_data[0][1])
                remetente = (msg.get("From") or "").lower()
                if not any(h in remetente for h in TWOFA_FROM_HINTS):
                    continue
                # só emails recebidos depois do início do login
                try:
                    dt = parsedate_to_datetime(msg.get("Date"))
                    if dt and dt.timestamp() < desde_utc.timestamp() - margem:
                        continue
                except Exception:
                    pass
                corpo = (msg.get("Subject") or "") + "\n" + _texto_email(msg)
                m = TWOFA_CODE_REGEX.search(corpo)
                if not m:
                    continue
                if any(k in corpo.lower() for k in TWOFA_KEYWORDS):
                    M.logout()
                    return m.group(1)  # email claramente de verificação
                if fallback is None:
                    fallback = m.group(1)
            M.logout()
        except Exception as e:
            log(f"Erro IMAP (retry): {e}")
        esperar(6, 9)
    if fallback:
        log("Sem email com palavra de verificação; a usar código de fallback.")
        return fallback
    raise Exception("Código 2FA não encontrado no email dentro do tempo limite.")

# ── Navegação + exportação ────────────────────────────────────────────────────
def descarregar_excel(driver, download_dir):
    log("A navegar para a Lista de Reservas...")
    # Tenta clicar diretamente em "Lista de Reservas"; se não estiver visível,
    # expande primeiro o menu "Reservas".
    if not clicar_por_texto(driver, "Lista de Reservas"):
        clicar_por_texto(driver, "Reservas")
        esperar(1.5, 2.5)
        clicar_por_texto(driver, "Lista de Reservas")
    esperar(2.5, 3.5)
    driver.save_screenshot("/tmp/tg_4_lista.png")

    log("A abrir menu 'Ações'...")
    if not clicar_por_texto(driver, "Ações"):
        clicar_por_texto(driver, "Acções")  # variação ortográfica
    esperar(1, 2)
    driver.save_screenshot("/tmp/tg_5_acoes.png")

    # A opção certa é "Exportar Reservas" (NÃO "Exportar Valores de Reservas").
    log("A clicar em 'Exportar Reservas'...")
    if not clicar_por_texto(driver, "Exportar Reservas"):
        driver.save_screenshot("/tmp/tg_debug.png")
        raise Exception("Opção 'Exportar Reservas' não encontrada. Ver screenshots de debug.")
    esperar(2, 3)
    driver.save_screenshot("/tmp/tg_6_exportar.png")

    log("A aguardar download do XLSX...")
    ficheiro = esperar_download(download_dir, timeout=90)
    log(f"XLSX descarregado: {ficheiro}")
    return ficheiro

def esperar_download(download_dir, timeout=60):
    deadline = time.time() + timeout
    while time.time() < deadline:
        # ignora downloads a meio (.crdownload)
        if not glob.glob(os.path.join(download_dir, "*.crdownload")):
            fichs = glob.glob(os.path.join(download_dir, "*.xlsx")) + \
                    glob.glob(os.path.join(download_dir, "*.xls"))
            if fichs:
                return sorted(fichs, key=os.path.getmtime)[-1]
        time.sleep(1.5)
    raise Exception("Download do Excel não concluído no tempo limite.")

# ── Parsing (mesma lógica verificada no index.html) ───────────────────────────
def tg_data(v):
    if v is None or v == "":
        return None
    if hasattr(v, "strftime"):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    if re.match(r"^\d{4}-\d{2}-\d{2}", s):
        return s[:10]
    if re.match(r"^\d{2}/\d{2}/\d{4}", s):
        p = re.split(r"[/\s]", s)
        return f"{p[2]}-{p[1]}-{p[0]}"
    return None

def tg_hora(v):
    if v is None or v == "":
        return ""
    if hasattr(v, "hour"):
        hh, mm = v.hour, v.minute
        return "" if (hh == 0 and mm == 0) else f"{hh:02d}:{mm:02d}"
    m = re.search(r"(\d{1,2}):(\d{2})", str(v))
    if m:
        hh, mm = int(m.group(1)), int(m.group(2))
        return "" if (hh == 0 and mm == 0) else f"{hh:02d}:{mm:02d}"
    return ""

def tg_num(v):
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(" ", "")
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except Exception:
        return 0.0

def tg_int(v):
    try:
        return int(round(tg_num(v)))
    except Exception:
        return 0

def tg_aloj(a):
    s = str(a or "").strip()
    return "Casa Santiago II Funchal" if s == "Casa Santiago II" else s

def tg_canal(c):
    s = str(c or "").strip()
    if s == "Booking.com":
        return "Booking"
    if s == "Airbnb":
        return "AirBnB"
    return s

def tg_estado(e):
    s = str(e or "").strip()
    if s in ("Cancelada", "Cancelled"):
        return "Cancelado"
    return s

def processar_excel(ficheiro):
    log(f"A processar Excel: {ficheiro}")
    wb = openpyxl.load_workbook(ficheiro, data_only=True)
    ws = wb[wb.sheetnames[0]]
    linhas = list(ws.iter_rows(values_only=True))
    if not linhas:
        raise Exception("Ficheiro Excel vazio.")

    headers = [str(c).strip() if c is not None else "" for c in linhas[0]]
    log(f"Colunas encontradas: {headers}")
    idx = {h: i for i, h in enumerate(headers)}

    def cel(row, nome):
        i = idx.get(nome)
        return row[i] if (i is not None and i < len(row)) else None

    reservas = []
    ignorados_bloqueio = 0
    for row in linhas[1:]:
        id_r    = str(cel(row, "Reserva") or "").strip()
        hospede = str(cel(row, "Hóspede") or "").strip()
        estado  = tg_estado(cel(row, "Estado"))
        canal   = tg_canal(cel(row, "Canal"))

        if estado == "Indisponivel" or canal == "Interno":
            ignorados_bloqueio += 1
            continue
        if not id_r or not hospede:
            continue

        total    = tg_num(cel(row, "Valor Reserva"))
        comissao = tg_num(cel(row, "Comissão Canal"))
        criancas = tg_int(cel(row, "Crianças sujeitas TMT")) + tg_int(cel(row, "Crianças não sujeitas TMT"))

        reservas.append({
            "id":                id_r,
            "hospede":           hospede,
            "checkin":           tg_data(cel(row, "Checkin")),
            "hora_checkin":      tg_hora(cel(row, "Hora Prevista Checkin")),
            "checkout":          tg_data(cel(row, "Checkout")),
            "hora_checkout":     tg_hora(cel(row, "Hora Prevista Checkout")),
            "noites":            tg_int(cel(row, "Noites")),
            "adultos":           tg_int(cel(row, "Adultos")),
            "criancas":          criancas,
            "bebes":             0,
            "telefone":          "",
            "email":             "",
            "pais":              "",
            "codigo_pais":       "",
            "alojamento":        tg_aloj(cel(row, "Alojamento")),
            "tmt":               0,
            "total":             total,
            "estado":            estado,
            "estado_pagamento":  "",
            "canal":             canal,
            "comissao":          comissao,
            "comissao_pct":      round(comissao / total * 100, 1) if total > 0 else 0,
            "id_canal":          id_r,
            "data_criacao":      tg_data(cel(row, "Reservado em")),
            "antecedencia":      0,
            "checkin_efetuado":  "",
            "checkout_efetuado": "",
            "notas_canal":       "",
            "fatura":            "",
            "estado_aima":       "",
        })

    log(f"{len(reservas)} reservas válidas ({ignorados_bloqueio} bloqueios internos ignorados).")
    for r in reservas[:3]:
        log(f"  ex: {r['id']} | {r['hospede']} | {r['alojamento']} | {r['canal']} | {r['checkin']}→{r['checkout']} | {r['total']}€")
    return reservas

# ── Importar via API (Cloudflare Worker) ──────────────────────────────────────
def importar_worker(reservas):
    if not reservas:
        raise Exception("Sem reservas para importar — abortado para não apagar dados.")
    log(f"A enviar {len(reservas)} reservas para a API...")
    resp = requests.post(
        API_URL.rstrip("/") + "/api/bot/import",
        headers={"Authorization": f"Bearer {BOT_SECRET}", "Content-Type": "application/json"},
        json={"reservas": reservas},
        timeout=120,
    )
    resp.raise_for_status()
    body = resp.json()
    log(f"✅ Importação concluída! {body.get('count', len(reservas))} reservas guardadas.")

# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    log("🤖 Bot Talkguest → Dashboard iniciado")
    driver, download_dir = criar_driver()
    try:
        fazer_login(driver)
        ficheiro = descarregar_excel(driver, download_dir)
        reservas = processar_excel(ficheiro)
        importar_worker(reservas)
    finally:
        driver.quit()
        log("Browser fechado.")

if __name__ == "__main__":
    main()
